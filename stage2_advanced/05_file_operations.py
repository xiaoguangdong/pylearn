# -*- coding: utf-8 -*-
"""
第二阶段：文件操作与数据序列化
演示文件读写、JSON/YAML处理、CSV/Excel处理和路径管理
"""

import json
import csv
from pathlib import Path


def demo_file_read_write():
    """演示文件读写"""
    print("=" * 50)
    print("文件读写演示")
    print("=" * 50)
    
    # 写入文本文件
    file_path = "demo.txt"
    content = "这是第一行\n这是第二行\n这是第三行"
    
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"已写入文件: {file_path}")
    
    # 读取文本文件
    with open(file_path, "r", encoding="utf-8") as f:
        read_content = f.read()
    print(f"读取内容:\n{read_content}")
    
    # 逐行读取
    print("\n逐行读取:")
    with open(file_path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f, 1):
            print(f"  第{i}行: {line.strip()}")
    
    # 清理
    Path(file_path).unlink()
    print()
    
    print()


def demo_json_operations():
    """演示JSON处理"""
    print("=" * 50)
    print("JSON处理演示")
    print("=" * 50)
    
    # Python对象转JSON
    data = {
        "姓名": "张三",
        "年龄": 25,
        "爱好": ["编程", "阅读", "运动"],
        "地址": {
            "城市": "北京",
            "街道": "中关村"
        }
    }
    
    json_file = "demo.json"
    
    # 写入JSON文件
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已写入JSON文件: {json_file}")
    
    # 读取JSON文件
    with open(json_file, "r", encoding="utf-8") as f:
        loaded_data = json.load(f)
    print(f"读取的JSON数据: {json.dumps(loaded_data, ensure_ascii=False, indent=2)}")
    
    # JSON字符串转换
    json_str = json.dumps(data, ensure_ascii=False)
    print(f"\nJSON字符串: {json_str[:50]}...")
    
    parsed_data = json.loads(json_str)
    print(f"解析后的数据姓名: {parsed_data['姓名']}")
    
    # 清理
    Path(json_file).unlink()
    
    print()


def demo_yaml_operations():
    """演示YAML处理（需要pyyaml库）"""
    print("=" * 50)
    print("YAML处理演示")
    print("=" * 50)
    
    try:
        import yaml
        
        data = {
            "姓名": "李四",
            "年龄": 30,
            "技能": ["Python", "JavaScript", "SQL"],
            "配置": {
                "主题": "dark",
                "语言": "zh-CN"
            }
        }
        
        yaml_file = "demo.yaml"
        
        # 写入YAML文件
        with open(yaml_file, "w", encoding="utf-8") as f:
            yaml.dump(data, f, allow_unicode=True, default_flow_style=False)
        print(f"已写入YAML文件: {yaml_file}")
        
        # 读取YAML文件
        with open(yaml_file, "r", encoding="utf-8") as f:
            loaded_data = yaml.safe_load(f)
        print(f"读取的YAML数据: {loaded_data}")
        
        # 清理
        Path(yaml_file).unlink()
        
    except ImportError:
        print("提示: 需要安装pyyaml库才能使用YAML功能")
        print("安装命令: pip install pyyaml")
    
    print()


def demo_csv_operations():
    """演示CSV文件处理"""
    print("=" * 50)
    print("CSV文件处理演示")
    print("=" * 50)
    
    csv_file = "demo.csv"
    
    # 写入CSV文件
    data = [
        ["姓名", "年龄", "城市"],
        ["张三", 25, "北京"],
        ["李四", 30, "上海"],
        ["王五", 28, "广州"]
    ]
    
    with open(csv_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(data)
    print(f"已写入CSV文件: {csv_file}")
    
    # 读取CSV文件
    print("\n读取CSV文件:")
    with open(csv_file, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            print(f"  {row}")
    
    # 使用字典方式读写CSV
    dict_data = [
        {"姓名": "赵六", "年龄": 32, "城市": "深圳"},
        {"姓名": "孙七", "年龄": 27, "城市": "杭州"}
    ]
    
    with open("demo_dict.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["姓名", "年龄", "城市"])
        writer.writeheader()
        writer.writerows(dict_data)
    print("\n已写入字典格式CSV文件")
    
    with open("demo_dict.csv", "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        print("读取字典格式CSV:")
        for row in reader:
            print(f"  {row}")
    
    # 清理
    Path(csv_file).unlink()
    Path("demo_dict.csv").unlink()
    
    print()


def demo_excel_operations():
    """演示Excel文件处理（需要openpyxl或pandas库）"""
    print("=" * 50)
    print("Excel文件处理演示")
    print("=" * 50)
    
    try:
        import openpyxl
        
        excel_file = "demo.xlsx"
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "学生信息"
        
        # 写入数据
        headers = ["姓名", "年龄", "成绩"]
        data = [
            ["张三", 20, 85],
            ["李四", 21, 92],
            ["王五", 19, 78]
        ]
        
        ws.append(headers)
        for row in data:
            ws.append(row)
        
        wb.save(excel_file)
        print(f"已写入Excel文件: {excel_file}")
        
        # 读取Excel文件
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        print("\n读取Excel文件:")
        for row in ws.iter_rows(values_only=True):
            print(f"  {row}")
        
        # 清理
        Path(excel_file).unlink()
        
    except ImportError:
        print("提示: 需要安装openpyxl库才能使用Excel功能")
        print("安装命令: pip install openpyxl")
        print("\n或者使用pandas处理Excel:")
        print("  pip install pandas openpyxl")
    
    print()


def demo_pathlib():
    """演示路径管理（pathlib）"""
    print("=" * 50)
    print("路径管理演示（pathlib）")
    print("=" * 50)
    
    # 创建Path对象
    current_dir = Path(".")
    print(f"当前目录: {current_dir.absolute()}")
    
    # 路径拼接
    file_path = current_dir / "data" / "test.txt"
    print(f"拼接路径: {file_path}")
    
    # 路径属性
    print(f"\n路径属性:")
    print(f"  父目录: {file_path.parent}")
    print(f"  文件名: {file_path.name}")
    print(f"  文件名（不含扩展名）: {file_path.stem}")
    print(f"  扩展名: {file_path.suffix}")
    
    # 路径操作
    test_file = Path("test_file.txt")
    test_file.write_text("测试内容", encoding="utf-8")
    print(f"\n文件是否存在: {test_file.exists()}")
    print(f"是否为文件: {test_file.is_file()}")
    print(f"是否为目录: {test_file.is_dir()}")
    
    # 创建目录
    test_dir = Path("test_dir")
    test_dir.mkdir(exist_ok=True)
    print(f"\n创建目录: {test_dir}")
    print(f"目录是否存在: {test_dir.exists()}")
    
    # 清理
    test_file.unlink()
    test_dir.rmdir()
    
    print()


if __name__ == "__main__":
    # 运行所有演示
    demo_file_read_write()
    demo_json_operations()
    demo_yaml_operations()
    demo_csv_operations()
    demo_excel_operations()
    demo_pathlib()
    
    print("=" * 50)
    print("演示完成！")
    print("=" * 50)

